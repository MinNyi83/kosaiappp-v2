#!/usr/bin/env python3
"""
KosAI Report Generator - Demo Version
Generates sample reports to demonstrate formatting and customization.
"""

import os
from datetime import datetime, timedelta

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing required packages...")
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter


# ============================================================================
# CUSTOMIZE YOUR STYLES HERE
# ============================================================================

class Styles:
    """Customizable report styles - modify these to change the look and feel."""

    # Color Palette
    COLORS = {
        'primary': '1F4E79',      # Deep blue
        'secondary': '2E75B6',    # Medium blue
        'accent': 'FFC000',       # Gold
        'success': '00B050',      # Green
        'warning': 'FF6600',      # Orange
        'danger': 'FF0000',       # Red
    }

    # Header style
    HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    HEADER_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')

    # Title style
    TITLE_FONT = Font(name='Calibri', size=16, bold=True, color='1F4E79')

    # Subtitle style
    SUBTITLE_FONT = Font(name='Calibri', size=12, bold=True, color='2E75B6')

    # Data style
    DATA_FONT = Font(name='Calibri', size=10)

    # Total/Summary style
    TOTAL_FONT = Font(name='Calibri', size=11, bold=True, color='1F4E79')
    TOTAL_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

    # Alternating row colors
    ALT_ROW_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')

    # Border
    THIN_BORDER = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7'),
    )

    # Alignment
    LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)
    CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
    RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')

    # Number formats
    USD_FORMAT = '$#,##0.00'
    MMK_FORMAT = '#,##0.00 "Ks"'
    PERCENT_FORMAT = '0.0%'
    NUMBER_FORMAT = '#,##0'


# ============================================================================
# SAMPLE DATA (Replace with real database queries)
# ============================================================================

def get_sample_inventory():
    """Sample inventory data."""
    return [
        {'item_code': 'CCTV-001', 'item_name': 'Hikvision DS-2CD2143G2-I', 'category': 'CCTV', 'stock_qty': 45, 'unit_price': 89.99, 'unit_price_mmk': 180000, 'batch_code': 'BATCH-2024-001'},
        {'item_code': 'CCTV-002', 'item_name': 'Dahua IPC-HDW3849HP', 'category': 'CCTV', 'stock_qty': 32, 'unit_price': 75.50, 'unit_price_mmk': 151000, 'batch_code': 'BATCH-2024-001'},
        {'item_code': 'NET-001', 'item_name': 'Ubiquiti UniFi Switch 24 PoE', 'category': 'Networking', 'stock_qty': 18, 'unit_price': 399.00, 'unit_price_mmk': 798000, 'batch_code': 'BATCH-2024-002'},
        {'item_code': 'NET-002', 'item_name': 'MikroTik hEX S', 'category': 'Networking', 'stock_qty': 3, 'unit_price': 69.00, 'unit_price_mmk': 138000, 'batch_code': 'BATCH-2024-002'},
        {'item_code': 'WIFI-001', 'item_name': 'Ubiquiti U6 Pro', 'category': 'WiFi', 'stock_qty': 25, 'unit_price': 149.00, 'unit_price_mmk': 298000, 'batch_code': 'BATCH-2024-003'},
        {'item_code': 'NAS-001', 'item_name': 'Synology DS923+', 'category': 'NAS', 'stock_qty': 8, 'unit_price': 549.99, 'unit_price_mmk': 1099980, 'batch_code': 'BATCH-2024-003'},
        {'item_code': 'CCTV-003', 'item_name': 'Axis P3245-V Dome', 'category': 'CCTV', 'stock_qty': 12, 'unit_price': 299.00, 'unit_price_mmk': 598000, 'batch_code': 'BATCH-2024-001'},
        {'item_code': 'NET-003', 'item_name': 'Cisco SG350-28', 'category': 'Networking', 'stock_qty': 2, 'unit_price': 199.00, 'unit_price_mmk': 398000, 'batch_code': 'BATCH-2024-002'},
        {'item_code': 'WIFI-002', 'item_name': 'TP-Link EAP670', 'category': 'WiFi', 'stock_qty': 4, 'unit_price': 119.00, 'unit_price_mmk': 238000, 'batch_code': 'BATCH-2024-003'},
        {'item_code': 'NAS-002', 'item_name': 'QNAP TS-464', 'category': 'NAS', 'stock_qty': 0, 'unit_price': 499.00, 'unit_price_mmk': 998000, 'batch_code': 'BATCH-2024-003'},
    ]

def get_sample_clients():
    """Sample client data."""
    return [
        {'company_name': 'Myanmar Tech Solutions', 'contact_person': 'U Aung Min', 'phone': '+95 9 123 456 789', 'amc_status': 'Active', 'amc_start': '2024-01-01', 'amc_end': '2025-01-01'},
        {'company_name': 'Yangon Business Center', 'contact_person': 'Daw Su Lin', 'phone': '+95 9 234 567 890', 'amc_status': 'Active', 'amc_start': '2024-03-15', 'amc_end': '2025-03-15'},
        {'company_name': 'Golden Palace Hotel', 'contact_person': 'U Kyaw Zin', 'phone': '+95 9 345 678 901', 'amc_status': 'Expired', 'amc_start': '2023-06-01', 'amc_end': '2024-06-01'},
        {'company_name': ' Mandalay Industrial Zone', 'contact_person': 'U Tun Aung', 'phone': '+95 9 456 789 012', 'amc_status': 'Active', 'amc_start': '2024-02-01', 'amc_end': '2025-02-01'},
        {'company_name': 'ABC Trading Co.', 'contact_person': 'Daw Mi Mi', 'phone': '+95 9 567 890 123', 'amc_status': 'Individual', 'amc_start': None, 'amc_end': None},
        {'company_name': 'Shwe Pharma Ltd', 'contact_person': 'U Soe Thein', 'phone': '+95 9 678 901 234', 'amc_status': 'Inactive', 'amc_start': '2023-09-01', 'amc_end': '2024-09-01'},
        {'company_name': 'Sky Net Services', 'contact_person': 'Daw Phyu Phyu', 'phone': '+95 9 789 012 345', 'amc_status': 'No AMC', 'amc_start': None, 'amc_end': None},
        {'company_name': 'Green Valley Resort', 'contact_person': 'U Htut Aung', 'phone': '+95 9 890 123 456', 'amc_status': 'Active', 'amc_start': '2024-04-01', 'amc_end': '2025-04-01'},
    ]

def get_sample_services():
    """Sample service records."""
    return [
        {'job_id': 'JOB-001', 'service_type': 'CCTV', 'status': 'Completed', 'client': 'Myanmar Tech Solutions', 'technician': 'U Win Aung', 'created_at': '2024-07-01', 'duration_days': 2},
        {'job_id': 'JOB-002', 'service_type': 'Networking', 'status': 'Completed', 'client': 'Yangon Business Center', 'technician': 'U Kyaw Swar', 'created_at': '2024-07-03', 'duration_days': 1},
        {'job_id': 'JOB-003', 'service_type': 'WiFi', 'status': 'Pending', 'client': 'Golden Palace Hotel', 'technician': 'U Myo Zaw', 'created_at': '2024-07-10', 'duration_days': None},
        {'job_id': 'JOB-004', 'service_type': 'CCTV', 'status': 'In Progress', 'client': 'Mandalay Industrial Zone', 'technician': 'U Win Aung', 'created_at': '2024-07-12', 'duration_days': None},
        {'job_id': 'JOB-005', 'service_type': 'NAS', 'status': 'Completed', 'client': 'ABC Trading Co.', 'technician': 'U Kyaw Swar', 'created_at': '2024-07-15', 'duration_days': 3},
        {'job_id': 'JOB-006', 'service_type': 'General Maintenance', 'status': 'Completed', 'client': 'Shwe Pharma Ltd', 'technician': 'U Myo Zaw', 'created_at': '2024-07-18', 'duration_days': 1},
        {'job_id': 'JOB-007', 'service_type': 'CCTV', 'status': 'Pending', 'client': 'Sky Net Services', 'technician': 'U Win Aung', 'created_at': '2024-07-20', 'duration_days': None},
        {'job_id': 'JOB-008', 'service_type': 'Networking', 'status': 'Completed', 'client': 'Green Valley Resort', 'technician': 'U Kyaw Swar', 'created_at': '2024-07-22', 'duration_days': 2},
    ]

def get_sample_technicians():
    """Sample technician data."""
    return [
        {'name': 'U Win Aung', 'role': 'Technician', 'total_jobs': 45, 'completed': 42, 'completion_rate': 93.3},
        {'name': 'U Kyaw Swar', 'role': 'Technician', 'total_jobs': 38, 'completed': 36, 'completion_rate': 94.7},
        {'name': 'U Myo Zaw', 'role': 'Technician', 'total_jobs': 32, 'completed': 28, 'completion_rate': 87.5},
        {'name': 'Daw Khin Myo', 'role': 'Sales', 'total_jobs': 15, 'completed': 15, 'completion_rate': 100.0},
    ]

def get_sample_cash_transactions():
    """Sample cash transactions."""
    return [
        {'date': '2024-07-01', 'type': 'Deposit', 'currency': 'USD', 'amount': 2500.00, 'exchange_rate': 2100, 'notes': 'Payment for CCTV installation - Myanmar Tech Solutions'},
        {'date': '2024-07-05', 'type': 'Deposit', 'currency': 'MMK', 'amount': 500000, 'exchange_rate': 2100, 'notes': 'AMC renewal fee - Yangon Business Center'},
        {'date': '2024-07-10', 'type': 'Withdrawal', 'currency': 'USD', 'amount': 500.00, 'exchange_rate': 2100, 'notes': 'Equipment purchase reimbursement'},
        {'date': '2024-07-15', 'type': 'Deposit', 'currency': 'USD', 'amount': 1800.00, 'exchange_rate': 2100, 'notes': 'Network setup - Mandalay Industrial Zone'},
        {'date': '2024-07-20', 'type': 'Withdrawal', 'currency': 'MMK', 'amount': 200000, 'exchange_rate': 2100, 'notes': 'Technician allowances'},
        {'date': '2024-07-25', 'type': 'Deposit', 'currency': 'USD', 'amount': 3200.00, 'exchange_rate': 2100, 'notes': 'Annual maintenance contract - Green Valley Resort'},
    ]


# ============================================================================
# REPORT GENERATOR
# ============================================================================

class ReportGenerator:
    """Generate and format reports."""

    def __init__(self):
        self.styles = Styles()

    def write_report(self, wb, title, subtitle, columns, data, summary=None, sheet_name=None):
        """Write a formatted report to worksheet."""
        if sheet_name is None:
            sheet_name = title[:31]

        ws = wb.create_sheet(title=sheet_name)

        # Title
        ws.merge_cells(f'A1:{get_column_letter(len(columns))}1')
        ws['A1'] = title
        ws['A1'].font = self.styles.TITLE_FONT
        ws['A1'].alignment = self.styles.LEFT_ALIGN

        # Subtitle
        ws.merge_cells(f'A2:{get_column_letter(len(columns))}2')
        ws['A2'] = subtitle
        ws['A2'].font = self.styles.SUBTITLE_FONT
        ws['A2'].alignment = self.styles.LEFT_ALIGN

        # Timestamp
        ws.merge_cells(f'A3:{get_column_letter(len(columns))}3')
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(name='Calibri', size=9, italic=True, color='666666')

        # Headers
        start_row = 5
        for col_idx, (key, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = self.styles.HEADER_FONT
            cell.fill = self.styles.HEADER_FILL
            cell.alignment = self.styles.CENTER_ALIGN
            cell.border = self.styles.THIN_BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Data rows
        for row_idx, row_data in enumerate(data):
            row_num = start_row + 1 + row_idx

            for col_idx, (key, header, width) in enumerate(columns, 1):
                value = row_data.get(key)
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = self.styles.DATA_FONT
                cell.alignment = self.styles.LEFT_ALIGN
                cell.border = self.styles.THIN_BORDER

                # Alternate row colors
                if row_idx % 2 == 1:
                    cell.fill = self.styles.ALT_ROW_FILL

                # Format numbers
                if 'price' in key or 'amount' in key or 'balance' in key:
                    cell.number_format = self.styles.USD_FORMAT
                elif 'mmk' in key.lower():
                    cell.number_format = self.styles.MMK_FORMAT
                elif 'rate' in key.lower():
                    cell.number_format = self.styles.PERCENT_FORMAT
                elif 'qty' in key.lower() or 'count' in key.lower():
                    cell.number_format = self.styles.NUMBER_FORMAT

        # Summary section
        if summary:
            summary_row = start_row + len(data) + 2
            ws.cell(row=summary_row, column=1, value="SUMMARY").font = self.styles.SUBTITLE_FONT

            for idx, (key, value) in enumerate(summary.items()):
                row = summary_row + 1 + idx
                label = key.replace('_', ' ').title()
                ws.cell(row=row, column=1, value=label).font = self.styles.TOTAL_FONT
                ws.cell(row=row, column=1).fill = self.styles.TOTAL_FILL

                val_cell = ws.cell(row=row, column=2, value=value)
                val_cell.font = self.styles.TOTAL_FONT
                val_cell.fill = self.styles.TOTAL_FILL
                val_cell.border = self.styles.THIN_BORDER

                if 'usd' in key.lower() or 'price' in key.lower():
                    val_cell.number_format = self.styles.USD_FORMAT
                elif 'mmk' in key.lower():
                    val_cell.number_format = self.styles.MMK_FORMAT
                elif 'rate' in key.lower():
                    val_cell.number_format = self.styles.PERCENT_FORMAT

        # Freeze panes
        ws.freeze_panes = f'A{start_row + 1}'

        # Auto-filter
        if data:
            last_col = get_column_letter(len(columns))
            ws.auto_filter.ref = f'A{start_row}:{last_col}{start_row + len(data)}'

        return ws


# ============================================================================
# MAIN - Generate All Reports
# ============================================================================

def main():
    """Generate all sample reports."""
    print("=" * 60)
    print("KosAI Report Generator - Demo")
    print("=" * 60)

    wb = Workbook()
    wb.remove(wb.active)

    generator = ReportGenerator()

    # 1. Inventory Summary
    print("\n[1/8] Generating Inventory Summary...")
    inventory = get_sample_inventory()
    for item in inventory:
        item['total_value_usd'] = item['stock_qty'] * item['unit_price']
        item['total_value_mmk'] = item['stock_qty'] * item['unit_price_mmk']

    generator.write_report(
        wb,
        title="Inventory Summary Report",
        subtitle="Stock levels, pricing, and total values",
        columns=[
            ('item_code', 'Item Code', 15),
            ('item_name', 'Item Name', 35),
            ('category', 'Category', 18),
            ('stock_qty', 'Quantity', 12),
            ('unit_price', 'Unit Price (USD)', 18),
            ('unit_price_mmk', 'Unit Price (MMK)', 18),
            ('total_value_usd', 'Total Value (USD)', 18),
            ('total_value_mmk', 'Total Value (MMK)', 18),
            ('batch_code', 'Batch Code', 18),
        ],
        data=inventory,
        summary={
            'total_items': len(inventory),
            'total_stock_qty': sum(i['stock_qty'] for i in inventory),
            'total_value_usd': sum(i['total_value_usd'] for i in inventory),
            'total_value_mmk': sum(i['total_value_mmk'] for i in inventory),
        }
    )

    # 2. Low Stock Alert
    print("[2/8] Generating Low Stock Alert...")
    low_stock = [i for i in inventory if i['stock_qty'] <= 5]
    generator.write_report(
        wb,
        title="Low Stock Alert",
        subtitle="Items requiring restocking attention",
        columns=[
            ('item_code', 'Item Code', 15),
            ('item_name', 'Item Name', 35),
            ('category', 'Category', 18),
            ('stock_qty', 'Current Stock', 15),
            ('unit_price', 'Unit Price (USD)', 18),
            ('total_value_usd', 'Value at Risk (USD)', 18),
        ],
        data=low_stock,
        summary={
            'total_low_stock_items': len(low_stock),
            'total_value_at_risk': sum(i['total_value_usd'] for i in low_stock),
        }
    )

    # 3. Client Summary
    print("[3/8] Generating Client Summary...")
    clients = get_sample_clients()
    generator.write_report(
        wb,
        title="Client Summary Report",
        subtitle="Complete client portfolio with AMC status",
        columns=[
            ('company_name', 'Company Name', 30),
            ('contact_person', 'Contact Person', 22),
            ('phone', 'Phone', 22),
            ('amc_status', 'AMC Status', 15),
            ('amc_start', 'AMC Start', 15),
            ('amc_end', 'AMC End', 15),
        ],
        data=clients,
        summary={
            'total_clients': len(clients),
            'active_amc': sum(1 for c in clients if c['amc_status'] == 'Active'),
            'expired_amc': sum(1 for c in clients if c['amc_status'] == 'Expired'),
        }
    )

    # 4. Service Records
    print("[4/8] Generating Service Records...")
    services = get_sample_services()
    generator.write_report(
        wb,
        title="Service Records Report",
        subtitle="All service jobs with status and duration",
        columns=[
            ('job_id', 'Job ID', 12),
            ('service_type', 'Service Type', 20),
            ('status', 'Status', 15),
            ('client', 'Client', 30),
            ('technician', 'Technician', 22),
            ('created_at', 'Created Date', 15),
            ('duration_days', 'Duration (Days)', 18),
        ],
        data=services,
        summary={
            'total_jobs': len(services),
            'completed': sum(1 for s in services if s['status'] == 'Completed'),
            'pending': sum(1 for s in services if s['status'] == 'Pending'),
            'in_progress': sum(1 for s in services if s['status'] == 'In Progress'),
        }
    )

    # 5. Technician Performance
    print("[5/8] Generating Technician Performance...")
    technicians = get_sample_technicians()
    generator.write_report(
        wb,
        title="Technician Performance Report",
        subtitle="Job completion rates by technician",
        columns=[
            ('name', 'Technician Name', 25),
            ('role', 'Role', 15),
            ('total_jobs', 'Total Assigned', 15),
            ('completed', 'Completed', 12),
            ('completion_rate', 'Completion Rate (%)', 18),
        ],
        data=technicians,
        summary={
            'avg_completion_rate': sum(t['completion_rate'] for t in technicians) / len(technicians),
        }
    )

    # 6. Cash Transactions
    print("[6/8] Generating Cash Transactions...")
    transactions = get_sample_cash_transactions()
    generator.write_report(
        wb,
        title="Cash Transactions Report",
        subtitle="All cash flow transactions",
        columns=[
            ('date', 'Date', 15),
            ('type', 'Type', 12),
            ('currency', 'Currency', 12),
            ('amount', 'Amount', 15),
            ('exchange_rate', 'Exchange Rate', 15),
            ('notes', 'Notes', 40),
        ],
        data=transactions,
        summary={
            'total_deposits_usd': sum(t['amount'] for t in transactions if t['type'] == 'Deposit' and t['currency'] == 'USD'),
            'total_withdrawals_usd': sum(t['amount'] for t in transactions if t['type'] == 'Withdrawal' and t['currency'] == 'USD'),
        }
    )

    # 7. Cash Safe Summary
    print("[7/8] Generating Cash Safe Summary...")
    safe_data = [
        {'safe_id': 1, 'usd_balance': 8500.00, 'mmk_balance': 2500000, 'last_updated': '2024-07-25'},
    ]
    generator.write_report(
        wb,
        title="Cash Safe Summary",
        subtitle="Current cash safe balances",
        columns=[
            ('safe_id', 'Safe ID', 10),
            ('usd_balance', 'USD Balance', 18),
            ('mmk_balance', 'MMK Balance', 18),
            ('last_updated', 'Last Updated', 15),
        ],
        data=safe_data,
        summary={
            'total_usd': sum(s['usd_balance'] for s in safe_data),
            'total_mmk': sum(s['mmk_balance'] for s in safe_data),
        }
    )

    # 8. Monthly Cash Flow
    print("[8/8] Generating Monthly Cash Flow...")
    monthly_flow = [
        {'month': '2024-07', 'usd_deposited': 7500.00, 'usd_withdrawn': 500.00, 'mmk_deposited': 500000, 'mmk_withdrawn': 200000},
    ]
    generator.write_report(
        wb,
        title="Monthly Cash Flow Report",
        subtitle="Cash flow trends over time",
        columns=[
            ('month', 'Month', 15),
            ('usd_deposited', 'USD Deposited', 18),
            ('usd_withdrawn', 'USD Withdrawn', 18),
            ('mmk_deposited', 'MMK Deposited', 18),
            ('mmk_withdrawn', 'MMK Withdrawn', 18),
        ],
        data=monthly_flow,
        summary={}
    )

    # Save workbook
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    output_file = f"KosAI_Reports_Demo_{timestamp}.xlsx"

    try:
        wb.save(output_file)
        print(f"\n{'=' * 60}")
        print(f"SUCCESS! Reports exported to: {output_file}")
        print(f"Total sheets: {len(wb.sheetnames)}")
        print(f"Sheets: {', '.join(wb.sheetnames)}")
        print(f"{'=' * 60}")
    except Exception as e:
        print(f"\n[ERROR] Failed to save workbook: {e}")


if __name__ == "__main__":
    main()
