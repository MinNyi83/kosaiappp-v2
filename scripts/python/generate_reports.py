#!/usr/bin/env python3
"""
KosAI Comprehensive Report Generator
Generates all possible reports from the database and exports to Excel.
Customizable styles and functions.
"""

import sqlite3
import os
from datetime import datetime, timedelta
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing required packages...")
    os.system("pip install openpyxl pandas")
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side, numbers
    )
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION - Customize your report styles here
# ============================================================================

class ReportConfig:
    """Central configuration for all report styles and settings."""

    # Color Palette
    COLORS = {
        'primary': '1F4E79',      # Deep blue
        'secondary': '2E75B6',    # Medium blue
        'accent': 'FFC000',       # Gold
        'success': '00B050',      # Green
        'warning': 'FF6600',      # Orange
        'danger': 'FF0000',       # Red
        'header_bg': '1F4E79',    # Header background
        'header_font': 'FFFFFF',  # Header font color
        'alt_row': 'D6E4F0',      # Alternating row color
        'total_bg': 'D9E2F3',     # Total row background
    }

    # Fonts
    FONTS = {
        'title': Font(name='Calibri', size=16, bold=True, color='1F4E79'),
        'subtitle': Font(name='Calibri', size=12, bold=True, color='2E75B6'),
        'header': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'data': Font(name='Calibri', size=10),
        'total': Font(name='Calibri', size=11, bold=True, color='1F4E79'),
        'currency': Font(name='Calibri', size=10, color='00B050'),
    }

    # Fills
    FILLS = {
        'header': PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid'),
        'alt_row': PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
        'total': PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid'),
        'success': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'warning': PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
        'danger': PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid'),
    }

    # Borders
    BORDER = Border(
        left=Side(style='thin', color='B4C6E7'),
        right=Side(style='thin', color='B4C6E7'),
        top=Side(style='thin', color='B4C6E7'),
        bottom=Side(style='thin', color='B4C6E7'),
    )

    # Alignment
    ALIGNMENT = {
        'left': Alignment(horizontal='left', vertical='center', wrap_text=True),
        'center': Alignment(horizontal='center', vertical='center'),
        'right': Alignment(horizontal='right', vertical='center'),
    }

    # Number Formats
    NUMBER_FORMATS = {
        'currency_usd': '$#,##0.00',
        'currency_mmk': '#,##0.00 "Ks"',
        'percent': '0.0%',
        'number': '#,##0',
        'date': 'YYYY-MM-DD',
        'datetime': 'YYYY-MM-DD HH:MM:SS',
    }


# ============================================================================
# DATABASE CONNECTION
# ============================================================================

class Database:
    """Database connection handler."""

    def __init__(self, db_path=None):
        if db_path is None:
            # Try to find the local D1 database
            possible_paths = [
                r'.wrangler/state/v3/d1/miniflare-D1DatabaseObject/1f6511a010eea77a87edada2cd2f8bd03d36571a83f5e8a8359f1d7e94856a92.sqlite',
                r'.wrangler/state/v3/d1/miniflare-D1DatabaseObject/*/sqlite.db',
            ]
            for pattern in possible_paths:
                import glob
                matches = glob.glob(pattern)
                if matches:
                    db_path = matches[0]
                    break

        if db_path is None or not os.path.exists(db_path):
            raise FileNotFoundError("Database not found. Please specify the correct path.")

        self.conn = sqlite3.connect(db_path)
        self.conn.row_factory = sqlite3.Row

    def query(self, sql, params=None):
        """Execute query and return results as list of dicts."""
        cursor = self.conn.cursor()
        if params:
            cursor.execute(sql, params)
        else:
            cursor.execute(sql)
        return [dict(row) for row in cursor.fetchall()]

    def close(self):
        self.conn.close()


# ============================================================================
# REPORT GENERATORS
# ============================================================================

class ReportGenerator:
    """Generate various reports from the database."""

    def __init__(self, db: Database, config: ReportConfig = None):
        self.db = db
        self.config = config or ReportConfig()

    # ------------------------------------------------------------------------
    # 1. INVENTORY REPORTS
    # ------------------------------------------------------------------------

    def inventory_summary(self):
        """Complete inventory summary with stock levels and values."""
        data = self.db.query("""
            SELECT
                item_code,
                item_name,
                category,
                stock_qty,
                unit_price,
                unit_price_mmk,
                buying_price,
                (stock_qty * unit_price) as total_value_usd,
                (stock_qty * unit_price_mmk) as total_value_mmk,
                batch_code
            FROM inventory_stock
            ORDER BY category, item_name
        """)
        return {
            'title': 'Inventory Summary Report',
            'subtitle': 'Stock levels, pricing, and total values',
            'data': data,
            'columns': [
                ('item_code', 'Item Code', 15),
                ('item_name', 'Item Name', 30),
                ('category', 'Category', 20),
                ('stock_qty', 'Quantity', 12),
                ('unit_price', 'Unit Price (USD)', 18),
                ('unit_price_mmk', 'Unit Price (MMK)', 18),
                ('buying_price', 'Buying Price', 15),
                ('total_value_usd', 'Total Value (USD)', 18),
                ('total_value_mmk', 'Total Value (MMK)', 18),
                ('batch_code', 'Batch Code', 15),
            ],
            'summary': {
                'total_items': len(data),
                'total_stock': sum(r['stock_qty'] or 0 for r in data),
                'total_value_usd': sum(r['total_value_usd'] or 0 for r in data),
                'total_value_mmk': sum(r['total_value_mmk'] or 0 for r in data),
            }
        }

    def inventory_by_category(self):
        """Inventory grouped by category."""
        data = self.db.query("""
            SELECT
                category,
                COUNT(*) as item_count,
                SUM(stock_qty) as total_qty,
                SUM(stock_qty * unit_price) as total_value_usd,
                SUM(stock_qty * unit_price_mmk) as total_value_mmk,
                AVG(unit_price) as avg_price_usd
            FROM inventory_stock
            GROUP BY category
            ORDER BY total_value_usd DESC
        """)
        return {
            'title': 'Inventory by Category',
            'subtitle': 'Stock distribution across categories',
            'data': data,
            'columns': [
                ('category', 'Category', 25),
                ('item_count', 'Item Count', 12),
                ('total_qty', 'Total Quantity', 15),
                ('total_value_usd', 'Total Value (USD)', 18),
                ('total_value_mmk', 'Total Value (MMK)', 18),
                ('avg_price_usd', 'Avg Price (USD)', 15),
            ],
            'summary': {
                'total_categories': len(data),
                'total_items': sum(r['item_count'] or 0 for r in data),
                'total_value_usd': sum(r['total_value_usd'] or 0 for r in data),
            }
        }

    def inventory_low_stock(self, threshold=5):
        """Items with low stock levels."""
        data = self.db.query(f"""
            SELECT
                item_code,
                item_name,
                category,
                stock_qty,
                unit_price,
                (stock_qty * unit_price) as value_usd,
                CASE
                    WHEN stock_qty = 0 THEN 'OUT OF STOCK'
                    WHEN stock_qty <= {threshold} THEN 'LOW'
                    ELSE 'OK'
                END as status
            FROM inventory_stock
            WHERE stock_qty <= {threshold}
            ORDER BY stock_qty ASC, item_name
        """)
        return {
            'title': f'Low Stock Alert (Threshold: {threshold})',
            'subtitle': 'Items requiring restocking attention',
            'data': data,
            'columns': [
                ('item_code', 'Item Code', 15),
                ('item_name', 'Item Name', 30),
                ('category', 'Category', 20),
                ('stock_qty', 'Current Stock', 15),
                ('unit_price', 'Unit Price (USD)', 15),
                ('value_usd', 'Value at Risk (USD)', 18),
                ('status', 'Status', 15),
            ],
            'summary': {
                'total_low_stock': len(data),
                'out_of_stock': sum(1 for r in data if r['status'] == 'OUT OF STOCK'),
                'total_value_at_risk': sum(r['value_usd'] or 0 for r in data),
            }
        }

    def inventory_by_batch(self):
        """Inventory grouped by batch code."""
        data = self.db.query("""
            SELECT
                s.batch_code,
                b.supplier,
                b.buying_price as batch_buying_price,
                b.created_at as batch_date,
                COUNT(*) as items_in_batch,
                SUM(s.stock_qty) as total_qty,
                SUM(s.stock_qty * s.unit_price) as total_value_usd
            FROM inventory_stock s
            LEFT JOIN inventory_batches b ON s.batch_code = b.batch_code
            GROUP BY s.batch_code
            ORDER BY b.created_at DESC
        """)
        return {
            'title': 'Inventory by Batch',
            'subtitle': 'Stock grouped by purchase batch',
            'data': data,
            'columns': [
                ('batch_code', 'Batch Code', 18),
                ('supplier', 'Supplier', 25),
                ('batch_buying_price', 'Batch Buying Price', 18),
                ('batch_date', 'Batch Date', 15),
                ('items_in_batch', 'Items in Batch', 15),
                ('total_qty', 'Total Quantity', 15),
                ('total_value_usd', 'Total Value (USD)', 18),
            ],
            'summary': {}
        }

    # ------------------------------------------------------------------------
    # 2. CLIENT REPORTS
    # ------------------------------------------------------------------------

    def client_summary(self):
        """Complete client list with AMC status."""
        data = self.db.query("""
            SELECT
                id,
                company_name,
                contact_person,
                phone,
                address,
                amc_status,
                amc_start,
                amc_end,
                CASE
                    WHEN amc_end < date('now') THEN 'Expired'
                    WHEN amc_end >= date('now') THEN 'Active'
                    ELSE 'No End Date'
                END as computed_status
            FROM clients
            ORDER BY company_name
        """)
        return {
            'title': 'Client Summary Report',
            'subtitle': 'Complete client portfolio with AMC status',
            'data': data,
            'columns': [
                ('id', 'Client ID', 15),
                ('company_name', 'Company Name', 30),
                ('contact_person', 'Contact Person', 25),
                ('phone', 'Phone', 18),
                ('address', 'Address', 35),
                ('amc_status', 'AMC Status', 15),
                ('amc_start', 'AMC Start', 15),
                ('amc_end', 'AMC End', 15),
                ('computed_status', 'Computed Status', 15),
            ],
            'summary': {
                'total_clients': len(data),
                'active_amc': sum(1 for r in data if r['amc_status'] == 'Active'),
                'expired_amc': sum(1 for r in data if r['amc_status'] == 'Expired'),
            }
        }

    def client_amc_breakdown(self):
        """AMC status breakdown."""
        data = self.db.query("""
            SELECT
                amc_status,
                COUNT(*) as client_count,
                GROUP_CONCAT(company_name, ', ') as companies
            FROM clients
            GROUP BY amc_status
            ORDER BY client_count DESC
        """)
        return {
            'title': 'AMC Status Breakdown',
            'subtitle': 'Client distribution by contract status',
            'data': data,
            'columns': [
                ('amc_status', 'AMC Status', 20),
                ('client_count', 'Client Count', 15),
                ('companies', 'Companies', 50),
            ],
            'summary': {
                'total_clients': sum(r['client_count'] or 0 for r in data),
            }
        }

    def client_service_history(self):
        """Client service history summary."""
        data = self.db.query("""
            SELECT
                c.company_name,
                c.amc_status,
                COUNT(sr.id) as total_services,
                SUM(CASE WHEN sr.status = 'Completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN sr.status = 'Pending' THEN 1 ELSE 0 END) as pending,
                SUM(CASE WHEN sr.status = 'In Progress' THEN 1 ELSE 0 END) as in_progress,
                MIN(sr.created_at) as first_service,
                MAX(sr.created_at) as last_service
            FROM clients c
            LEFT JOIN service_records sr ON c.id = sr.client_id
            GROUP BY c.id, c.company_name
            ORDER BY total_services DESC
        """)
        return {
            'title': 'Client Service History',
            'subtitle': 'Service records per client',
            'data': data,
            'columns': [
                ('company_name', 'Company Name', 30),
                ('amc_status', 'AMC Status', 15),
                ('total_services', 'Total Services', 15),
                ('completed', 'Completed', 12),
                ('pending', 'Pending', 12),
                ('in_progress', 'In Progress', 12),
                ('first_service', 'First Service', 15),
                ('last_service', 'Last Service', 15),
            ],
            'summary': {
                'total_clients': len(data),
                'total_services': sum(r['total_services'] or 0 for r in data),
            }
        }

    # ------------------------------------------------------------------------
    # 3. SERVICE / JOB REPORTS
    # ------------------------------------------------------------------------

    def service_summary(self):
        """Service records summary."""
        data = self.db.query("""
            SELECT
                sr.id,
                sr.service_type,
                sr.status,
                sr.job_description,
                c.company_name,
                t.name as technician_name,
                sr.created_at,
                sr.completion_time,
                CASE
                    WHEN sr.completion_time IS NOT NULL AND sr.arrival_time IS NOT NULL
                    THEN JULIANDAY(sr.completion_time) - JULIANDAY(sr.arrival_time)
                    ELSE NULL
                END as duration_days
            FROM service_records sr
            LEFT JOIN clients c ON sr.client_id = c.id
            LEFT JOIN technicians t ON sr.technician_id = t.id
            ORDER BY sr.created_at DESC
        """)
        return {
            'title': 'Service Records Report',
            'subtitle': 'All service jobs with status and duration',
            'data': data,
            'columns': [
                ('id', 'Job ID', 15),
                ('service_type', 'Service Type', 18),
                ('status', 'Status', 15),
                ('job_description', 'Description', 35),
                ('company_name', 'Client', 25),
                ('technician_name', 'Technician', 20),
                ('created_at', 'Created', 15),
                ('completion_time', 'Completed', 15),
                ('duration_days', 'Duration (Days)', 15),
            ],
            'summary': {
                'total_jobs': len(data),
                'completed': sum(1 for r in data if r['status'] == 'Completed'),
                'pending': sum(1 for r in data if r['status'] == 'Pending'),
                'in_progress': sum(1 for r in data if r['status'] == 'In Progress'),
            }
        }

    def service_by_type(self):
        """Service records grouped by service type."""
        data = self.db.query("""
            SELECT
                service_type,
                COUNT(*) as total_jobs,
                SUM(CASE WHEN status = 'Completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending,
                SUM(CASE WHEN status = 'In Progress' THEN 1 ELSE 0 END) as in_progress,
                SUM(CASE WHEN status = 'Cancelled' THEN 1 ELSE 0 END) as cancelled,
                ROUND(AVG(CASE
                    WHEN completion_time IS NOT NULL AND arrival_time IS NOT NULL
                    THEN JULIANDAY(completion_time) - JULIANDAY(arrival_time)
                END), 1) as avg_duration_days
            FROM service_records
            GROUP BY service_type
            ORDER BY total_jobs DESC
        """)
        return {
            'title': 'Services by Type',
            'subtitle': 'Job distribution across service categories',
            'data': data,
            'columns': [
                ('service_type', 'Service Type', 20),
                ('total_jobs', 'Total Jobs', 12),
                ('completed', 'Completed', 12),
                ('pending', 'Pending', 12),
                ('in_progress', 'In Progress', 12),
                ('cancelled', 'Cancelled', 12),
                ('avg_duration_days', 'Avg Duration (Days)', 18),
            ],
            'summary': {
                'total_jobs': sum(r['total_jobs'] or 0 for r in data),
            }
        }

    def service_by_status(self):
        """Service records grouped by status."""
        data = self.db.query("""
            SELECT
                status,
                COUNT(*) as job_count,
                GROUP_CONCAT(DISTINCT service_type) as service_types
            FROM service_records
            GROUP BY status
            ORDER BY job_count DESC
        """)
        return {
            'title': 'Services by Status',
            'subtitle': 'Job distribution by current status',
            'data': data,
            'columns': [
                ('status', 'Status', 18),
                ('job_count', 'Job Count', 12),
                ('service_types', 'Service Types', 40),
            ],
            'summary': {
                'total_jobs': sum(r['job_count'] or 0 for r in data),
            }
        }

    def service_monthly_trend(self):
        """Monthly service trend."""
        data = self.db.query("""
            SELECT
                strftime('%Y-%m', created_at) as month,
                COUNT(*) as total_jobs,
                SUM(CASE WHEN status = 'Completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending
            FROM service_records
            GROUP BY strftime('%Y-%m', created_at)
            ORDER BY month DESC
            LIMIT 12
        """)
        return {
            'title': 'Monthly Service Trend',
            'subtitle': 'Job volume over time (last 12 months)',
            'data': data,
            'columns': [
                ('month', 'Month', 15),
                ('total_jobs', 'Total Jobs', 12),
                ('completed', 'Completed', 12),
                ('pending', 'Pending', 12),
            ],
            'summary': {}
        }

    # ------------------------------------------------------------------------
    # 4. TECHNICIAN REPORTS
    # ------------------------------------------------------------------------

    def technician_summary(self):
        """Technician list with roles and activity."""
        data = self.db.query("""
            SELECT
                t.id,
                t.name,
                t.role,
                t.phone,
                t.email,
                t.active,
                t.last_login,
                COUNT(sr.id) as total_jobs,
                SUM(CASE WHEN sr.status = 'Completed' THEN 1 ELSE 0 END) as completed_jobs
            FROM technicians t
            LEFT JOIN service_records sr ON t.id = sr.technician_id
            GROUP BY t.id, t.name
            ORDER BY total_jobs DESC
        """)
        return {
            'title': 'Technician Summary Report',
            'subtitle': 'Technician profiles and job performance',
            'data': data,
            'columns': [
                ('id', 'Technician ID', 15),
                ('name', 'Name', 25),
                ('role', 'Role', 15),
                ('phone', 'Phone', 18),
                ('email', 'Email', 25),
                ('active', 'Active', 10),
                ('last_login', 'Last Login', 18),
                ('total_jobs', 'Total Jobs', 12),
                ('completed_jobs', 'Completed Jobs', 15),
            ],
            'summary': {
                'total_technicians': len(data),
                'active_technicians': sum(1 for r in data if r['active'] == 1),
            }
        }

    def technician_performance(self):
        """Technician performance metrics."""
        data = self.db.query("""
            SELECT
                t.name as technician_name,
                t.role,
                COUNT(sr.id) as total_assigned,
                SUM(CASE WHEN sr.status = 'Completed' THEN 1 ELSE 0 END) as completed,
                ROUND(
                    CASE
                        WHEN COUNT(sr.id) > 0
                        THEN (SUM(CASE WHEN sr.status = 'Completed' THEN 1 ELSE 0 END) * 100.0 / COUNT(sr.id))
                        ELSE 0
                    END, 1
                ) as completion_rate
            FROM technicians t
            LEFT JOIN service_records sr ON t.id = sr.technician_id
            GROUP BY t.id, t.name
            ORDER BY completion_rate DESC
        """)
        return {
            'title': 'Technician Performance Report',
            'subtitle': 'Job completion rates by technician',
            'data': data,
            'columns': [
                ('technician_name', 'Technician', 25),
                ('role', 'Role', 15),
                ('total_assigned', 'Total Assigned', 15),
                ('completed', 'Completed', 12),
                ('completion_rate', 'Completion Rate (%)', 18),
            ],
            'summary': {
                'avg_completion_rate': sum(r['completion_rate'] or 0 for r in data) / len(data) if data else 0,
            }
        }

    def technician_workload(self):
        """Current workload per technician."""
        data = self.db.query("""
            SELECT
                t.name as technician_name,
                SUM(CASE WHEN sr.status = 'Pending' THEN 1 ELSE 0 END) as pending_jobs,
                SUM(CASE WHEN sr.status = 'In Progress' THEN 1 ELSE 0 END) as in_progress_jobs,
                SUM(CASE WHEN sr.status = 'Completed' THEN 1 ELSE 0 END) as completed_jobs,
                COUNT(sr.id) as total_jobs
            FROM technicians t
            LEFT JOIN service_records sr ON t.id = sr.technician_id
            GROUP BY t.id, t.name
            ORDER BY (pending_jobs + in_progress_jobs) DESC
        """)
        return {
            'title': 'Technician Workload Report',
            'subtitle': 'Current job allocation across technicians',
            'data': data,
            'columns': [
                ('technician_name', 'Technician', 25),
                ('pending_jobs', 'Pending', 12),
                ('in_progress_jobs', 'In Progress', 15),
                ('completed_jobs', 'Completed', 12),
                ('total_jobs', 'Total Jobs', 12),
            ],
            'summary': {}
        }

    # ------------------------------------------------------------------------
    # 5. FINANCIAL REPORTS
    # ------------------------------------------------------------------------

    def cash_safe_summary(self):
        """Cash safe balances."""
        data = self.db.query("""
            SELECT
                id,
                usd_balance,
                mmk_balance,
                (usd_balance + mmk_balance / 3500) as estimated_total_usd
            FROM cash_safes
        """)
        return {
            'title': 'Cash Safe Summary',
            'subtitle': 'Current cash safe balances',
            'data': data,
            'columns': [
                ('id', 'Safe ID', 10),
                ('usd_balance', 'USD Balance', 18),
                ('mmk_balance', 'MMK Balance', 18),
                ('estimated_total_usd', 'Est. Total (USD)', 18),
            ],
            'summary': {
                'total_usd': sum(r['usd_balance'] or 0 for r in data),
                'total_mmk': sum(r['mmk_balance'] or 0 for r in data),
            }
        }

    def cash_transactions(self):
        """Cash transaction history."""
        data = self.db.query("""
            SELECT
                ct.id,
                ct.transaction_type,
                ct.primary_currency,
                ct.amount,
                ct.exchange_rate,
                ct.equivalent_amount,
                ct.notes,
                ct.created_at,
                ct.receive_mmk,
                sr.job_description
            FROM cash_transactions ct
            LEFT JOIN service_records sr ON ct.job_id = sr.id
            ORDER BY ct.created_at DESC
        """)
        return {
            'title': 'Cash Transactions Report',
            'subtitle': 'All cash flow transactions',
            'data': data,
            'columns': [
                ('id', 'Txn ID', 8),
                ('transaction_type', 'Type', 15),
                ('primary_currency', 'Currency', 12),
                ('amount', 'Amount', 15),
                ('exchange_rate', 'Exchange Rate', 15),
                ('equivalent_amount', 'Equivalent Amount', 18),
                ('notes', 'Notes', 30),
                ('created_at', 'Date', 18),
                ('job_description', 'Related Job', 25),
            ],
            'summary': {
                'total_deposits': sum(r['amount'] or 0 for r in data if r['transaction_type'] == 'Deposit'),
                'total_withdrawals': sum(r['amount'] or 0 for r in data if r['transaction_type'] == 'Withdrawal'),
            }
        }

    def cash_flow_monthly(self):
        """Monthly cash flow summary."""
        data = self.db.query("""
            SELECT
                strftime('%Y-%m', created_at) as month,
                SUM(CASE WHEN transaction_type = 'Deposit' AND primary_currency = 'USD' THEN amount ELSE 0 END) as usd_deposits,
                SUM(CASE WHEN transaction_type = 'Withdrawal' AND primary_currency = 'USD' THEN amount ELSE 0 END) as usd_withdrawals,
                SUM(CASE WHEN transaction_type = 'Deposit' AND primary_currency = 'MMK' THEN amount ELSE 0 END) as mmk_deposits,
                SUM(CASE WHEN transaction_type = 'Withdrawal' AND primary_currency = 'MMK' THEN amount ELSE 0 END) as mmk_withdrawals
            FROM cash_transactions
            GROUP BY strftime('%Y-%m', created_at)
            ORDER BY month DESC
            LIMIT 12
        """)
        return {
            'title': 'Monthly Cash Flow',
            'subtitle': 'Cash flow trends over time',
            'data': data,
            'columns': [
                ('month', 'Month', 15),
                ('usd_deposits', 'USD Deposits', 18),
                ('usd_withdrawals', 'USD Withdrawals', 18),
                ('mmk_deposits', 'MMK Deposits', 18),
                ('mmk_withdrawals', 'MMK Withdrawals', 18),
            ],
            'summary': {}
        }

    def service_fees(self):
        """Service fee schedule."""
        data = self.db.query("""
            SELECT
                id,
                service_type,
                fee_amount,
                currency,
                description
            FROM service_fees
            ORDER BY service_type, currency
        """)
        return {
            'title': 'Service Fees Schedule',
            'subtitle': 'Standard pricing for service types',
            'data': data,
            'columns': [
                ('id', 'ID', 8),
                ('service_type', 'Service Type', 25),
                ('fee_amount', 'Fee Amount', 15),
                ('currency', 'Currency', 12),
                ('description', 'Description', 40),
            ],
            'summary': {}
        }

    # ------------------------------------------------------------------------
    # 6. WARRANTY & RMA REPORTS
    # ------------------------------------------------------------------------

    def warranty_status(self):
        """Warranty status of installed devices."""
        data = self.db.query("""
            SELECT
                i.serial_number,
                i.device_name,
                c.company_name,
                i.installed_date,
                i.warranty_months,
                i.status,
                i.distributor,
                DATE(i.installed_date, '+' || i.warranty_months || ' months') as warranty_end,
                CASE
                    WHEN DATE(i.installed_date, '+' || i.warranty_months || ' months') >= date('now')
                    THEN 'Under Warranty'
                    ELSE 'Warranty Expired'
                END as warranty_status
            FROM inventory_items i
            LEFT JOIN clients c ON i.client_id = c.id
            ORDER BY warranty_end DESC
        """)
        return {
            'title': 'Warranty Status Report',
            'subtitle': 'Device warranty tracking',
            'data': data,
            'columns': [
                ('serial_number', 'Serial Number', 20),
                ('device_name', 'Device Name', 25),
                ('company_name', 'Client', 25),
                ('installed_date', 'Installed Date', 15),
                ('warranty_months', 'Warranty (Months)', 18),
                ('warranty_end', 'Warranty End', 15),
                ('status', 'Device Status', 15),
                ('warranty_status', 'Warranty Status', 18),
            ],
            'summary': {
                'total_devices': len(data),
                'under_warranty': sum(1 for r in data if r['warranty_status'] == 'Under Warranty'),
                'expired_warranty': sum(1 for r in data if r['warranty_status'] == 'Warranty Expired'),
            }
        }

    def rma_tracking(self):
        """RMA (Return Merchandise Authorization) tracking."""
        data = self.db.query("""
            SELECT
                i.serial_number,
                i.device_name,
                c.company_name,
                i.status,
                i.rma_tracking_id,
                i.distributor,
                i.installed_date
            FROM inventory_items i
            LEFT JOIN clients c ON i.client_id = c.id
            WHERE i.status IN ('Defective', 'RMA Sent', 'RMA Completed', 'Replaced')
            ORDER BY i.status, i.installed_date
        """)
        return {
            'title': 'RMA Tracking Report',
            'subtitle': 'Devices under RMA process',
            'data': data,
            'columns': [
                ('serial_number', 'Serial Number', 20),
                ('device_name', 'Device Name', 25),
                ('company_name', 'Client', 25),
                ('status', 'Status', 15),
                ('rma_tracking_id', 'RMA Tracking ID', 20),
                ('distributor', 'Distributor', 20),
                ('installed_date', 'Installed Date', 15),
            ],
            'summary': {
                'total_rma': len(data),
                'defective': sum(1 for r in data if r['status'] == 'Defective'),
                'rma_sent': sum(1 for r in data if r['status'] == 'RMA Sent'),
                'rma_completed': sum(1 for r in data if r['status'] == 'RMA Completed'),
            }
        }

    def device_installation_log(self):
        """Device installation log."""
        data = self.db.query("""
            SELECT
                i.serial_number,
                i.device_name,
                c.company_name,
                c.contact_person,
                i.installed_date,
                i.warranty_months,
                i.status,
                i.distributor,
                i.batch_code
            FROM inventory_items i
            LEFT JOIN clients c ON i.client_id = c.id
            ORDER BY i.installed_date DESC
        """)
        return {
            'title': 'Device Installation Log',
            'subtitle': 'Complete installation history',
            'data': data,
            'columns': [
                ('serial_number', 'Serial Number', 20),
                ('device_name', 'Device Name', 25),
                ('company_name', 'Client', 25),
                ('contact_person', 'Contact', 20),
                ('installed_date', 'Installation Date', 18),
                ('warranty_months', 'Warranty (Months)', 18),
                ('status', 'Status', 15),
                ('distributor', 'Distributor', 20),
                ('batch_code', 'Batch Code', 15),
            ],
            'summary': {
                'total_installations': len(data),
            }
        }

    # ------------------------------------------------------------------------
    # 8. ADDITIONAL REPORTS
    # ------------------------------------------------------------------------

    def monthly_job_trend(self):
        """Monthly job trend report."""
        data = self.db.query("""
            SELECT
                strftime('%Y-%m', created_at) as month,
                COUNT(*) as total_jobs,
                SUM(CASE WHEN status = 'Completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending,
                SUM(CASE WHEN status = 'In Progress' THEN 1 ELSE 0 END) as in_progress,
                ROUND(AVG(CASE WHEN status = 'Completed' THEN 1.0 ELSE 0.0 END) * 100, 1) as completion_rate
            FROM service_records
            GROUP BY strftime('%Y-%m', created_at)
            ORDER BY month DESC
            LIMIT 12
        """)
        return {
            'title': 'Monthly Job Trend',
            'subtitle': 'Job volume and completion rates over time',
            'data': data,
            'columns': [
                ('month', 'Month', 15),
                ('total_jobs', 'Total Jobs', 12),
                ('completed', 'Completed', 12),
                ('pending', 'Pending', 12),
                ('in_progress', 'In Progress', 12),
                ('completion_rate', 'Completion Rate (%)', 18),
            ],
            'summary': {
                'total_months': len(data),
            }
        }

    def client_revenue_report(self):
        """Client-wise job summary."""
        data = self.db.query("""
            SELECT
                c.company_name,
                c.amc_status,
                COUNT(sr.id) as total_jobs,
                SUM(CASE WHEN sr.status = 'Completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN sr.status = 'Pending' THEN 1 ELSE 0 END) as pending,
                SUM(CASE WHEN sr.status = 'In Progress' THEN 1 ELSE 0 END) as in_progress,
                MIN(sr.created_at) as first_job,
                MAX(sr.created_at) as last_job
            FROM clients c
            LEFT JOIN service_records sr ON c.id = sr.client_id
            GROUP BY c.id, c.company_name
            ORDER BY total_jobs DESC
        """)
        return {
            'title': 'Client-wise Job Summary',
            'subtitle': 'Job distribution across clients',
            'data': data,
            'columns': [
                ('company_name', 'Company', 30),
                ('amc_status', 'AMC Status', 15),
                ('total_jobs', 'Total Jobs', 12),
                ('completed', 'Completed', 12),
                ('pending', 'Pending', 12),
                ('in_progress', 'In Progress', 12),
                ('first_job', 'First Job', 15),
                ('last_job', 'Last Job', 15),
            ],
            'summary': {
                'total_clients': len(data),
                'total_jobs': sum(r['total_jobs'] or 0 for r in data),
            }
        }

    def technician_workload_detail(self):
        """Detailed technician workload."""
        data = self.db.query("""
            SELECT
                t.name as technician_name,
                t.role,
                SUM(CASE WHEN sr.status = 'Pending' THEN 1 ELSE 0 END) as pending,
                SUM(CASE WHEN sr.status = 'In Progress' THEN 1 ELSE 0 END) as in_progress,
                SUM(CASE WHEN sr.status = 'Completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN sr.status = 'Cancelled' THEN 1 ELSE 0 END) as cancelled,
                COUNT(sr.id) as total,
                ROUND(AVG(CASE WHEN sr.status = 'Completed' THEN 1.0 ELSE 0.0 END) * 100, 1) as completion_rate
            FROM technicians t
            LEFT JOIN service_records sr ON t.id = sr.technician_id
            GROUP BY t.id, t.name
            ORDER BY total DESC
        """)
        return {
            'title': 'Technician Workload Detail',
            'subtitle': 'Detailed workload breakdown by technician',
            'data': data,
            'columns': [
                ('technician_name', 'Technician', 25),
                ('role', 'Role', 15),
                ('pending', 'Pending', 10),
                ('in_progress', 'In Progress', 12),
                ('completed', 'Completed', 12),
                ('cancelled', 'Cancelled', 12),
                ('total', 'Total', 10),
                ('completion_rate', 'Completion Rate (%)', 18),
            ],
            'summary': {
                'total_technicians': len(data),
            }
        }

    def inventory_value_by_category(self):
        """Inventory value analysis by category."""
        data = self.db.query("""
            SELECT
                category,
                COUNT(*) as item_count,
                SUM(stock_qty) as total_qty,
                SUM(stock_qty * unit_price) as total_value_usd,
                SUM(stock_qty * unit_price_mmk) as total_value_mmk,
                ROUND(SUM(stock_qty * unit_price) * 100.0 /
                    (SELECT SUM(stock_qty * unit_price) FROM inventory_stock WHERE stock_qty * unit_price > 0), 1) as pct_of_total
            FROM inventory_stock
            WHERE stock_qty > 0
            GROUP BY category
            ORDER BY total_value_usd DESC
        """)
        return {
            'title': 'Inventory Value by Category',
            'subtitle': 'Stock value distribution across categories',
            'data': data,
            'columns': [
                ('category', 'Category', 25),
                ('item_count', 'Items', 10),
                ('total_qty', 'Total Qty', 12),
                ('total_value_usd', 'Value (USD)', 18),
                ('total_value_mmk', 'Value (MMK)', 18),
                ('pct_of_total', '% of Total', 12),
            ],
            'summary': {
                'total_categories': len(data),
                'total_value_usd': sum(r['total_value_usd'] or 0 for r in data),
                'total_value_mmk': sum(r['total_value_mmk'] or 0 for r in data),
            }
        }

    def warranty_expiry_alert(self):
        """Warranty expiry alert - devices expiring soon."""
        data = self.db.query("""
            SELECT
                i.serial_number,
                i.device_name,
                c.company_name,
                i.installed_date,
                i.warranty_months,
                DATE(i.installed_date, '+' || i.warranty_months || ' months') as warranty_end,
                CAST((JULIANDAY(DATE(i.installed_date, '+' || i.warranty_months || ' months')) - JULIANDAY('now')) AS INTEGER) as days_until_expiry,
                CASE
                    WHEN DATE(i.installed_date, '+' || i.warranty_months || ' months') < date('now') THEN 'EXPIRED'
                    WHEN DATE(i.installed_date, '+' || i.warranty_months || ' months') <= date('now', '+30 days') THEN 'EXPIRING SOON'
                    ELSE 'OK'
                END as status
            FROM inventory_items i
            LEFT JOIN clients c ON i.client_id = c.id
            WHERE DATE(i.installed_date, '+' || i.warranty_months || ' months') <= date('now', '+60 days')
            ORDER BY days_until_expiry ASC
        """)
        return {
            'title': 'Warranty Expiry Alert',
            'subtitle': 'Devices with warranty expiring within 60 days',
            'data': data,
            'columns': [
                ('serial_number', 'Serial Number', 20),
                ('device_name', 'Device Name', 25),
                ('company_name', 'Client', 25),
                ('installed_date', 'Installed', 15),
                ('warranty_end', 'Warranty End', 15),
                ('days_until_expiry', 'Days Left', 12),
                ('status', 'Status', 15),
            ],
            'summary': {
                'expiring_soon': sum(1 for r in data if r['status'] == 'EXPIRING SOON'),
                'already_expired': sum(1 for r in data if r['status'] == 'EXPIRED'),
            }
        }

    def amc_expiry_alert(self):
        """AMC expiry alert - contracts expiring soon."""
        data = self.db.query("""
            SELECT
                company_name,
                contact_person,
                phone,
                amc_status,
                amc_start,
                amc_end,
                CAST((JULIANDAY(amc_end) - JULIANDAY('now')) AS INTEGER) as days_until_expiry,
                CASE
                    WHEN amc_end < date('now') THEN 'EXPIRED'
                    WHEN amc_end <= date('now', '+30 days') THEN 'EXPIRING SOON'
                    ELSE 'OK'
                END as status
            FROM clients
            WHERE amc_status = 'Active'
                AND amc_end <= date('now', '+60 days')
            ORDER BY days_until_expiry ASC
        """)
        return {
            'title': 'AMC Expiry Alert',
            'subtitle': 'Active AMC contracts expiring within 60 days',
            'data': data,
            'columns': [
                ('company_name', 'Company', 30),
                ('contact_person', 'Contact', 20),
                ('phone', 'Phone', 18),
                ('amc_end', 'AMC End', 15),
                ('days_until_expiry', 'Days Left', 12),
                ('status', 'Status', 15),
            ],
            'summary': {
                'expiring_soon': sum(1 for r in data if r['status'] == 'EXPIRING SOON'),
                'already_expired': sum(1 for r in data if r['status'] == 'EXPIRED'),
            }
        }

    def job_duration_analysis(self):
        """Job duration analysis for completed jobs."""
        data = self.db.query("""
            SELECT
                sr.id as job_id,
                sr.service_type,
                c.company_name,
                t.name as technician_name,
                sr.arrival_time,
                sr.completion_time,
                ROUND((JULIANDAY(sr.completion_time) - JULIANDAY(sr.arrival_time)) * 24, 1) as duration_hours
            FROM service_records sr
            LEFT JOIN clients c ON sr.client_id = c.id
            LEFT JOIN technicians t ON sr.technician_id = t.id
            WHERE sr.status = 'Completed'
                AND sr.arrival_time IS NOT NULL
                AND sr.completion_time IS NOT NULL
            ORDER BY duration_hours DESC
        """)
        return {
            'title': 'Job Duration Analysis',
            'subtitle': 'Time taken for completed jobs',
            'data': data,
            'columns': [
                ('job_id', 'Job ID', 12),
                ('service_type', 'Type', 18),
                ('company_name', 'Client', 25),
                ('technician_name', 'Technician', 20),
                ('arrival_time', 'Arrival', 18),
                ('completion_time', 'Completion', 18),
                ('duration_hours', 'Duration (Hours)', 15),
            ],
            'summary': {
                'total_jobs': len(data),
                'avg_duration': round(sum(r['duration_hours'] or 0 for r in data) / len(data), 1) if data else 0,
                'max_duration': max((r['duration_hours'] or 0 for r in data), default=0),
            }
        }

    def device_status_overview(self):
        """Device status overview."""
        data = self.db.query("""
            SELECT
                status,
                COUNT(*) as count,
                GROUP_CONCAT(device_name, ', ') as sample_devices
            FROM inventory_items
            GROUP BY status
            ORDER BY count DESC
        """)
        return {
            'title': 'Device Status Overview',
            'subtitle': 'Distribution of device statuses',
            'data': data,
            'columns': [
                ('status', 'Status', 20),
                ('count', 'Count', 12),
                ('sample_devices', 'Sample Devices', 50),
            ],
            'summary': {
                'total_devices': sum(r['count'] or 0 for r in data),
                'active': next((r['count'] for r in data if r['status'] == 'Active'), 0),
                'defective': next((r['count'] for r in data if r['status'] == 'Defective'), 0),
            }
        }

    def batch_purchase_report(self):
        """Batch purchase report."""
        data = self.db.query("""
            SELECT
                b.batch_code,
                b.supplier,
                b.buying_price,
                b.created_at,
                COUNT(DISTINCT s.item_code) as items_in_batch,
                SUM(s.stock_qty) as total_qty,
                SUM(s.stock_qty * s.unit_price) as total_value_usd
            FROM inventory_batches b
            LEFT JOIN inventory_stock s ON b.batch_code = s.batch_code
            GROUP BY b.batch_code
            ORDER BY b.created_at DESC
        """)
        return {
            'title': 'Batch Purchase Report',
            'subtitle': 'Purchase batch details and stock allocation',
            'data': data,
            'columns': [
                ('batch_code', 'Batch Code', 18),
                ('supplier', 'Supplier', 25),
                ('buying_price', 'Buying Price', 15),
                ('created_at', 'Purchase Date', 15),
                ('items_in_batch', 'Items', 10),
                ('total_qty', 'Total Qty', 12),
                ('total_value_usd', 'Value (USD)', 18),
            ],
            'summary': {
                'total_batches': len(data),
                'total_value_usd': sum(r['total_value_usd'] or 0 for r in data),
            }
        }

    # ------------------------------------------------------------------------
    # 7. OPERATIONAL DASHBOARD
    # ------------------------------------------------------------------------

    def operational_dashboard(self):
        """High-level operational metrics."""
        # Get counts
        clients = self.db.query("SELECT COUNT(*) as cnt FROM clients")
        technicians = self.db.query("SELECT COUNT(*) as cnt FROM technicians WHERE active = 1")
        jobs = self.db.query("""
            SELECT
                COUNT(*) as total,
                SUM(CASE WHEN status = 'Completed' THEN 1 ELSE 0 END) as completed,
                SUM(CASE WHEN status = 'Pending' THEN 1 ELSE 0 END) as pending,
                SUM(CASE WHEN status = 'In Progress' THEN 1 ELSE 0 END) as in_progress
            FROM service_records
        """)
        inventory = self.db.query("SELECT COUNT(*) as cnt, SUM(stock_qty) as total_qty FROM inventory_stock")
        safe = self.db.query("SELECT SUM(usd_balance) as usd, SUM(mmk_balance) as mmk FROM cash_safes")

        data = [
            {'metric': 'Total Clients', 'value': clients[0]['cnt'], 'category': 'Clients'},
            {'metric': 'Active Technicians', 'value': technicians[0]['cnt'], 'category': 'Technicians'},
            {'metric': 'Total Jobs', 'value': jobs[0]['total'], 'category': 'Jobs'},
            {'metric': 'Completed Jobs', 'value': jobs[0]['completed'], 'category': 'Jobs'},
            {'metric': 'Pending Jobs', 'value': jobs[0]['pending'], 'category': 'Jobs'},
            {'metric': 'In Progress Jobs', 'value': jobs[0]['in_progress'], 'category': 'Jobs'},
            {'metric': 'Inventory Items', 'value': inventory[0]['cnt'], 'category': 'Inventory'},
            {'metric': 'Total Stock Quantity', 'value': inventory[0]['total_qty'], 'category': 'Inventory'},
            {'metric': 'USD Cash Safe', 'value': safe[0]['usd'] or 0, 'category': 'Finance'},
            {'metric': 'MMK Cash Safe', 'value': safe[0]['mmk'] or 0, 'category': 'Finance'},
        ]
        return {
            'title': 'Operational Dashboard',
            'subtitle': 'High-level business metrics',
            'data': data,
            'columns': [
                ('metric', 'Metric', 25),
                ('value', 'Value', 18),
                ('category', 'Category', 18),
            ],
            'summary': {}
        }


# ============================================================================
# EXCEL WRITER
# ============================================================================

class ExcelWriter:
    """Write reports to Excel with professional formatting."""

    def __init__(self, config: ReportConfig = None):
        self.config = config or ReportConfig()

    def write_report(self, wb: Workbook, report_data: dict, sheet_name: str = None):
        """Write a single report to a worksheet."""
        if sheet_name is None:
            # Excel sheet name max 31 chars, remove invalid characters
            invalid_chars = ':/\\?*[]'
            sheet_name = report_data['title'][:31]
            for char in invalid_chars:
                sheet_name = sheet_name.replace(char, '')

        ws = wb.create_sheet(title=sheet_name)
        config = self.config

        # Title
        ws.merge_cells('A1:Z1')
        ws['A1'] = report_data['title']
        ws['A1'].font = config.FONTS['title']
        ws['A1'].alignment = config.ALIGNMENT['left']

        # Subtitle
        ws.merge_cells('A2:Z2')
        ws['A2'] = report_data['subtitle']
        ws['A2'].font = config.FONTS['subtitle']
        ws['A2'].alignment = config.ALIGNMENT['left']

        # Timestamp
        ws.merge_cells('A3:Z3')
        ws['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A3'].font = Font(name='Calibri', size=9, italic=True, color='666666')

        # Headers (row 5)
        start_row = 5
        columns = report_data['columns']

        for col_idx, (key, header, width) in enumerate(columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = config.FONTS['header']
            cell.fill = config.FILLS['header']
            cell.alignment = config.ALIGNMENT['center']
            cell.border = config.BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Data rows
        data = report_data['data']
        for row_idx, row_data in enumerate(data):
            row_num = start_row + 1 + row_idx

            for col_idx, (key, header, width) in enumerate(columns, 1):
                value = row_data.get(key)
                cell = ws.cell(row=row_num, column=col_idx, value=value)
                cell.font = config.FONTS['data']
                cell.alignment = config.ALIGNMENT['left']
                cell.border = config.BORDER

                # Alternate row colors
                if row_idx % 2 == 1:
                    cell.fill = config.FILLS['alt_row']

                # Format currency columns
                if 'usd' in key.lower() or 'price' in key.lower() or 'balance' in key.lower() or 'amount' in key.lower():
                    cell.number_format = config.NUMBER_FORMATS['currency_usd']
                elif 'mmk' in key.lower():
                    cell.number_format = config.NUMBER_FORMATS['currency_mmk']
                elif 'rate' in key.lower() and 'exchange' not in key.lower():
                    cell.number_format = config.NUMBER_FORMATS['percent']
                elif 'qty' in key.lower() or 'count' in key.lower() or 'total' in key.lower():
                    cell.number_format = config.NUMBER_FORMATS['number']

        # Summary section
        summary = report_data.get('summary', {})
        if summary:
            summary_row = start_row + len(data) + 2
            ws.cell(row=summary_row, column=1, value="SUMMARY").font = config.FONTS['subtitle']

            for idx, (key, value) in enumerate(summary.items()):
                row = summary_row + 1 + idx
                label = key.replace('_', ' ').title()
                ws.cell(row=row, column=1, value=label).font = config.FONTS['total']
                ws.cell(row=row, column=1).fill = config.FILLS['total']

                val_cell = ws.cell(row=row, column=2, value=value)
                val_cell.font = config.FONTS['total']
                val_cell.fill = config.FILLS['total']
                val_cell.border = config.BORDER

                if 'usd' in key.lower() or 'price' in key.lower() or 'value' in key.lower():
                    val_cell.number_format = config.NUMBER_FORMATS['currency_usd']
                elif 'mmk' in key.lower():
                    val_cell.number_format = config.NUMBER_FORMATS['currency_mmk']
                elif 'rate' in key.lower():
                    val_cell.number_format = config.NUMBER_FORMATS['percent']

        # Freeze panes
        ws.freeze_panes = f'A{start_row + 1}'

        # Auto-filter
        if data:
            last_col = get_column_letter(len(columns))
            ws.auto_filter.ref = f'A{start_row}:{last_col}{start_row + len(data)}'

        return ws

    def add_chart(self, ws: Workbook, chart_type: str, data_range: str, title: str, anchor: str = 'A1'):
        """Add a chart to the worksheet."""
        if chart_type == 'bar':
            chart = BarChart()
        elif chart_type == 'pie':
            chart = PieChart()
        else:
            return

        chart.title = title
        chart.style = 10
        chart.y_axis.title = 'Value'
        chart.x_axis.title = 'Category'

        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        ws.add_chart(chart, anchor)


# ============================================================================
# MAIN EXECUTION
# ============================================================================

def main():
    """Generate all reports and export to Excel."""
    print("=" * 60)
    print("KosAI Comprehensive Report Generator")
    print("=" * 60)

    # Initialize database
    try:
        db = Database()
        print("[OK] Database connected")
    except FileNotFoundError as e:
        print(f"[ERROR] {e}")
        print("Please ensure the local D1 database exists.")
        return

    # Initialize report generator
    generator = ReportGenerator(db)

    # Create workbook
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    # Initialize Excel writer
    writer = ExcelWriter()

    # Generate all reports
    reports = [
        # Inventory Reports
        ("Inventory Summary", generator.inventory_summary),
        ("Inventory by Category", generator.inventory_by_category),
        ("Inventory Value by Category", generator.inventory_value_by_category),
        ("Low Stock Alert", generator.inventory_low_stock),
        ("Inventory by Batch", generator.inventory_by_batch),
        ("Batch Purchase Report", generator.batch_purchase_report),

        # Client Reports
        ("Client Summary", generator.client_summary),
        ("AMC Breakdown", generator.client_amc_breakdown),
        ("Client Service History", generator.client_service_history),
        ("Client-wise Job Summary", generator.client_revenue_report),
        ("AMC Expiry Alert", generator.amc_expiry_alert),

        # Service Reports
        ("Service Summary", generator.service_summary),
        ("Services by Type", generator.service_by_type),
        ("Services by Status", generator.service_by_status),
        ("Monthly Service Trend", generator.service_monthly_trend),
        ("Monthly Job Trend", generator.monthly_job_trend),
        ("Job Duration Analysis", generator.job_duration_analysis),

        # Technician Reports
        ("Technician Summary", generator.technician_summary),
        ("Technician Performance", generator.technician_performance),
        ("Technician Workload", generator.technician_workload),
        ("Technician Workload Detail", generator.technician_workload_detail),

        # Financial Reports
        ("Cash Safe Summary", generator.cash_safe_summary),
        ("Cash Transactions", generator.cash_transactions),
        ("Monthly Cash Flow", generator.cash_flow_monthly),
        ("Service Fees", generator.service_fees),

        # Warranty & RMA Reports
        ("Warranty Status", generator.warranty_status),
        ("Warranty Expiry Alert", generator.warranty_expiry_alert),
        ("RMA Tracking", generator.rma_tracking),
        ("Device Installation Log", generator.device_installation_log),
        ("Device Status Overview", generator.device_status_overview),

        # Dashboard
        ("Operational Dashboard", generator.operational_dashboard),
    ]

    print(f"\nGenerating {len(reports)} reports...")

    for name, generator_func in reports:
        try:
            report_data = generator_func()
            writer.write_report(wb, report_data)
            print(f"  [OK] {name}: {len(report_data['data'])} records")
        except Exception as e:
            print(f"  [ERROR] {name}: {e}")

    # Save workbook
    timestamp = datetime.now().strftime('%Y-%m-%d_%H%M%S')
    output_file = f"KosAI_Reports_{timestamp}.xlsx"

    try:
        wb.save(output_file)
        print(f"\n{'=' * 60}")
        print(f"SUCCESS! Reports exported to: {output_file}")
        print(f"Total sheets: {len(wb.sheetnames)}")
        print(f"{'=' * 60}")
    except Exception as e:
        print(f"\n[ERROR] Failed to save workbook: {e}")

    db.close()


if __name__ == "__main__":
    main()
